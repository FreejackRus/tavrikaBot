import os
import logging
import pandas as pd
from typing import Dict, Any, Tuple


RU_ACCOUNT_MAIN = "Главная касса"
RU_ACCOUNT_TRADES = "Торговые кассы"

ACCOUNT_NAME_MAP = {
    "Main cash register": RU_ACCOUNT_MAIN,
    "Trade cash registers": RU_ACCOUNT_TRADES,
    RU_ACCOUNT_MAIN: RU_ACCOUNT_MAIN,
    RU_ACCOUNT_TRADES: RU_ACCOUNT_TRADES,
}

CATEGORY_RU_MAP = {
    # Common English -> Russian mappings
    "Internal transfer": "Внутреннее перемещение",
    "Invoices payment": "Оплата счетов",
    "Sales": "Выручка",
    "Revenue": "Выручка",
    "Loan": "Займ",
    "Loans": "Займ",
    # Already Russian values pass-through
    "Внутреннее перемещение": "Внутреннее перемещение",
    "Выручка": "Выручка",
}


def _normalize_accounts(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    if "Account.Name" not in df.columns:
        # Nothing to normalize
        df["AccountNorm"] = None
        return df[df["AccountNorm"].isin({RU_ACCOUNT_MAIN, RU_ACCOUNT_TRADES})]

    def map_name(val: Any) -> Any:
        try:
            s = str(val or "").strip().lower()
        except Exception:
            s = ""
        if not s:
            return None
        # Robust matching for Russian and English variants, ignoring suffixes/prefixes
        if ("главная касс" in s) or ("main cash register" in s):
            return RU_ACCOUNT_MAIN
        if ("торгов" in s) or ("trade cash register" in s) or ("trade cash registers" in s):
            return RU_ACCOUNT_TRADES
        # Fallback to direct map if exact key exists
        return ACCOUNT_NAME_MAP.get(val)

    df["AccountNorm"] = df["Account.Name"].map(map_name)

    # Diagnostics: log unique names and mapping coverage
    try:
        src_names = sorted(set(str(x) for x in df["Account.Name"].dropna().unique().tolist()))
        norm_names = sorted(set(str(x) for x in df["AccountNorm"].dropna().unique().tolist()))
        logging.info("[cashflow] Account.Name uniques: %s", src_names)
        logging.info("[cashflow] AccountNorm uniques: %s", norm_names)
        logging.info("[cashflow] rows before norm: %d, after norm: %d", len(df), len(df[df["AccountNorm"].isin({RU_ACCOUNT_MAIN, RU_ACCOUNT_TRADES})]))
    except Exception:
        pass

    df = df[df["AccountNorm"].isin({RU_ACCOUNT_MAIN, RU_ACCOUNT_TRADES})]
    return df


def build_cashflow_tables(raw_json: Dict[str, Any]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    raw = raw_json.get("data", [])
    df = pd.json_normalize(raw)

    df = _normalize_accounts(df)

    # Ensure required columns exist
    for col in [
        "Sum.Incoming",
        "Sum.Outgoing",
        "FinalBalance.Money",
        "StartBalance.Money",
        "CashFlowCategory.HierarchyLevel1",
        "CashFlowCategory.HierarchyLevel2",
        "CashFlowCategory.HierarchyLevel3",
        "CashFlowCategory.Type",
    ]:
        if col not in df.columns:
            df[col] = 0

    cat_col = "CashFlowCategory.HierarchyLevel1"
    type_col = "CashFlowCategory.Type"

    # Отдельно выделим строки балансов (без категории) и движения (с категорией)
    balances_df = df[df[cat_col].isna()] if df[cat_col].isna().any() else df[df[cat_col] == None]
    # Include all flow types (OPERATIONAL and FINANCE)
    flows_mask = (df[cat_col].notna() if df[cat_col].notna().any() else df[cat_col] != None)
    flows_df = df[flows_mask]

    # 1) Остаток на начало (из строк без категории: StartBalance.Money)
    start_bal = (
        balances_df.groupby("AccountNorm")["StartBalance.Money"].sum().reindex([RU_ACCOUNT_MAIN, RU_ACCOUNT_TRADES]).fillna(0)
    )
    start_row = pd.DataFrame([start_bal.values], index=["Остаток на начало"], columns=[RU_ACCOUNT_MAIN, RU_ACCOUNT_TRADES])

    # 2) Обороты по категориям и типам: приход и расход отдельно
    incoming = (
        flows_df.groupby([type_col, cat_col, "AccountNorm"])  # type: ignore
        ["Sum.Incoming"].sum(min_count=1)
        .unstack(fill_value=0)
    )
    outgoing = (
        flows_df.groupby([type_col, cat_col, "AccountNorm"])  # type: ignore
        ["Sum.Outgoing"].sum(min_count=1)
        .unstack(fill_value=0)
    )

    # Map categories to Russian labels and aggregate duplicates
    if not incoming.empty:
        incoming.index = incoming.index.map(lambda x: (x[0], CATEGORY_RU_MAP.get(x[1], str(x[1]))))
        incoming = incoming.groupby(level=[0, 1]).sum()
    if not outgoing.empty:
        outgoing.index = outgoing.index.map(lambda x: (x[0], CATEGORY_RU_MAP.get(x[1], str(x[1]))))
        outgoing = outgoing.groupby(level=[0, 1]).sum()

    # Ensure columns exist
    for frame in (incoming, outgoing):
        for col in [RU_ACCOUNT_MAIN, RU_ACCOUNT_TRADES]:
            if col not in frame.columns:
                frame[col] = 0

    # 3) Остаток на конец (из строк без категории: FinalBalance.Money)
    end_bal = (
        balances_df.groupby("AccountNorm")["FinalBalance.Money"].sum().reindex([RU_ACCOUNT_MAIN, RU_ACCOUNT_TRADES]).fillna(0)
    )
    end_row = pd.DataFrame([end_bal.values], index=["Остаток на конец"], columns=[RU_ACCOUNT_MAIN, RU_ACCOUNT_TRADES])

    # 4) Итоговая таблица в нужном порядке
    rows = []
    # Начальные остатки
    rows.append({
        "Тип статьи": "Остаток на начало",
        "Статья ДДС 1-го уровня": "",
        "Статья ДДС 2-го уровня": "",
        "Статья ДДС 3-го уровня": "",
        RU_ACCOUNT_MAIN: float(start_bal.get(RU_ACCOUNT_MAIN, 0)),
        RU_ACCOUNT_TRADES: float(start_bal.get(RU_ACCOUNT_TRADES, 0)),
    })

    # Движения по типам и категориям
    for flow_type in ["OPERATIONAL", "FINANCE"]:
        type_label = "Операционная деятельность" if flow_type == "OPERATIONAL" else "Финансовая деятельность"
        
        # Фильтруем категории для текущего типа
        type_incoming = incoming[incoming.index.get_level_values(0) == flow_type] if not incoming.empty else pd.DataFrame()
        type_outgoing = outgoing[outgoing.index.get_level_values(0) == flow_type] if not outgoing.empty else pd.DataFrame()
        
        if not type_incoming.empty or not type_outgoing.empty:
            # Заголовок раздела
            rows.append({
                "Тип статьи": type_label,
                "Статья ДДС 1-го уровня": "",
                "Статья ДДС 2-го уровня": "",
                "Статья ДДС 3-го уровня": "",
                RU_ACCOUNT_MAIN: 0,
                RU_ACCOUNT_TRADES: 0,
            })

            # Категории текущего типа
            # Собираем множество категорий корректно, избегая конкатенации списков с приоритетом
            cats_in = list(type_incoming.index.get_level_values(1)) if not type_incoming.empty else []
            cats_out = list(type_outgoing.index.get_level_values(1)) if not type_outgoing.empty else []
            categories = sorted(set(cats_in) | set(cats_out))

            for cat in categories:
                # Приход
                if not type_incoming.empty and cat in type_incoming.index.get_level_values(1):
                    cat_incoming = type_incoming.xs((flow_type, cat), level=[0, 1])
                    rows.append({
                        "Тип статьи": type_label,
                        "Статья ДДС 1-го уровня": cat,
                        "Статья ДДС 2-го уровня": "",
                        "Статья ДДС 3-го уровня": "",
                        RU_ACCOUNT_MAIN: float(cat_incoming.get(RU_ACCOUNT_MAIN, 0)),
                        RU_ACCOUNT_TRADES: float(cat_incoming.get(RU_ACCOUNT_TRADES, 0)),
                    })

                # Расход (отрицательные значения)
                if not type_outgoing.empty and cat in type_outgoing.index.get_level_values(1):
                    cat_outgoing = type_outgoing.xs((flow_type, cat), level=[0, 1])
                    rows.append({
                        "Тип статьи": type_label,
                        "Статья ДДС 1-го уровня": cat,
                        "Статья ДДС 2-го уровня": "",
                        "Статья ДДС 3-го уровня": "",
                        RU_ACCOUNT_MAIN: -float(cat_outgoing.get(RU_ACCOUNT_MAIN, 0)),
                        RU_ACCOUNT_TRADES: -float(cat_outgoing.get(RU_ACCOUNT_TRADES, 0)),
                    })

    # Конечные остатки
    rows.append({
        "Тип статьи": "Остаток на конец",
        "Статья ДДС 1-го уровня": "",
        "Статья ДДС 2-го уровня": "",
        "Статья ДДС 3-го уровня": "",
        RU_ACCOUNT_MAIN: float(end_bal.get(RU_ACCOUNT_MAIN, 0)),
        RU_ACCOUNT_TRADES: float(end_bal.get(RU_ACCOUNT_TRADES, 0)),
    })

    # Создаем DataFrame с нужными колонками
    result = pd.DataFrame(rows, columns=[
        "Тип статьи",
        "Статья ДДС 1-го уровня",
        "Статья ДДС 2-го уровня",
        "Статья ДДС 3-го уровня",
        RU_ACCOUNT_MAIN,
        RU_ACCOUNT_TRADES,
    ])

    # Add total column
    result["Итого"] = result[[RU_ACCOUNT_MAIN, RU_ACCOUNT_TRADES]].sum(axis=1)

    return result, df


def build_cashflow_detailed_table(raw_json: Dict[str, Any]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    raw = raw_json.get("data", [])
    df = pd.json_normalize(raw)

    df = _normalize_accounts(df)

    # Ensure required columns exist
    for col in [
        "Sum.Incoming",
        "Sum.Outgoing",
        "FinalBalance.Money",
        "StartBalance.Money",
        "CashFlowCategory.HierarchyLevel1",
        "CashFlowCategory.HierarchyLevel2",
        "CashFlowCategory.HierarchyLevel3",
        "CashFlowCategory.Type",
    ]:
        if col not in df.columns:
            df[col] = 0

    cat_col = "CashFlowCategory.HierarchyLevel1"
    type_col = "CashFlowCategory.Type"

    # Балансы без категории
    # NA masking: use isna()/notna() consistently, avoiding element-wise comparisons to None
    balances_df = df[df[cat_col].isna()]
    # Движения по периоду: включаем все типы (OPERATIONAL и FINANCE)
    flows_mask = df[cat_col].notna()
    flows_df = df[flows_mask]

    # Остаток на начало
    start_bal = (
        balances_df.groupby("AccountNorm")["StartBalance.Money"].sum().reindex([RU_ACCOUNT_MAIN, RU_ACCOUNT_TRADES]).fillna(0)
    )

    # Разбивка по категориям: приход и расход отдельно
    incoming = (
        flows_df.groupby([type_col, cat_col, "AccountNorm"])  # type: ignore
        ["Sum.Incoming"].sum(min_count=1)
        .unstack(fill_value=0)
    )
    outgoing = (
        flows_df.groupby([type_col, cat_col, "AccountNorm"])  # type: ignore
        ["Sum.Outgoing"].sum(min_count=1)
        .unstack(fill_value=0)
    )

    # Normalize categories to Russian (2-й уровень индекса)
    if not incoming.empty:
        incoming.index = incoming.index.map(lambda x: (x[0], CATEGORY_RU_MAP.get(x[1], str(x[1]))))
        incoming = incoming.groupby(level=[0, 1]).sum()
    if not outgoing.empty:
        outgoing.index = outgoing.index.map(lambda x: (x[0], CATEGORY_RU_MAP.get(x[1], str(x[1]))))
        outgoing = outgoing.groupby(level=[0, 1]).sum()

    # Ensure columns exist
    for frame in (incoming, outgoing):
        for col in [RU_ACCOUNT_MAIN, RU_ACCOUNT_TRADES]:
            if col not in frame.columns:
                frame[col] = 0

    def _get_vals_by_type(frame: pd.DataFrame, flow_type: str, cat: str) -> Tuple[float, float]:
        if frame.empty or (flow_type, cat) not in frame.index:
            return (0.0, 0.0)
        row = frame.loc[(flow_type, cat)]
        return float(row.get(RU_ACCOUNT_MAIN, 0) or 0), float(row.get(RU_ACCOUNT_TRADES, 0) or 0)

    # Собираем строки детальной таблицы
    rows = []

    def add_row(type_label: str, l1: str = "", l2: str = "", l3: str = "", inout: str = "", main_val: float = 0.0, trade_val: float = 0.0):
        total = (main_val or 0) + (trade_val or 0)
        rows.append({
            "Тип статьи": type_label,
            "Статья ДДС 1-го уровня": l1,
            "Статья ДДС 2-ого уровня": l2,
            "Статья ДДС 3-ого уровня": l3,
            "Приход/Расход": inout,
            RU_ACCOUNT_MAIN: main_val,
            RU_ACCOUNT_TRADES: trade_val,
            "Итого": total,
        })

    # Row: Остаток на начало
    add_row("Остаток на начало", main_val=float(start_bal.get(RU_ACCOUNT_MAIN, 0)), trade_val=float(start_bal.get(RU_ACCOUNT_TRADES, 0)))

    # Движения по типам и категориям за период: включаем обе категории деятельности
    for flow_type in ["OPERATIONAL", "FINANCE"]:
        type_label = "Операционная деятельность" if flow_type == "OPERATIONAL" else "Финансовая деятельность"

        type_incoming = incoming[incoming.index.get_level_values(0) == flow_type] if not incoming.empty else pd.DataFrame()
        type_outgoing = outgoing[outgoing.index.get_level_values(0) == flow_type] if not outgoing.empty else pd.DataFrame()

        if not type_incoming.empty or not type_outgoing.empty:
            # Заголовок/раздел
            add_row(type_label)

            # Категории текущего типа
            cats_in = list(type_incoming.index.get_level_values(1)) if not type_incoming.empty else []
            cats_out = list(type_outgoing.index.get_level_values(1)) if not type_outgoing.empty else []
            categories = sorted(set(cats_in) | set(cats_out))

            for cat in categories:
                # Приход
                if not type_incoming.empty and cat in type_incoming.index.get_level_values(1):
                    in_main, in_trades = _get_vals_by_type(incoming, flow_type, cat)
                    add_row(type_label, cat, inout="Приход", main_val=in_main, trade_val=in_trades)

                # Расход (отображаем отрицательным)
                if not type_outgoing.empty and cat in type_outgoing.index.get_level_values(1):
                    out_main, out_trades = _get_vals_by_type(outgoing, flow_type, cat)
                    add_row(type_label, cat, inout="Расход", main_val=-out_main, trade_val=-out_trades)

    # Остаток на конец
    end_bal = (
        balances_df.groupby("AccountNorm")["FinalBalance.Money"].sum().reindex([RU_ACCOUNT_MAIN, RU_ACCOUNT_TRADES]).fillna(0)
    )
    add_row("Остаток на конец", main_val=float(end_bal.get(RU_ACCOUNT_MAIN, 0)), trade_val=float(end_bal.get(RU_ACCOUNT_TRADES, 0)))

    detailed = pd.DataFrame(rows, columns=[
        "Тип статьи",
        "Статья ДДС 1-го уровня",
        "Статья ДДС 2-ого уровня",
        "Статья ДДС 3-ого уровня",
        "Приход/Расход",
        RU_ACCOUNT_MAIN,
        RU_ACCOUNT_TRADES,
        "Итого",
    ])

    return detailed, df


def calculate_daily_movement(current: pd.DataFrame, previous: pd.DataFrame, account: str, type_col: str, cat_col: str) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Calculate incoming and outgoing movements for the day by comparing current and previous data."""
    # Current day totals
    curr_incoming = (
        current.groupby([type_col, cat_col, "AccountNorm"])["Sum.Incoming"]  # type: ignore
        .sum(min_count=1)
        .unstack(fill_value=0)
    )
    curr_outgoing = (
        current.groupby([type_col, cat_col, "AccountNorm"])["Sum.Outgoing"]  # type: ignore
        .sum(min_count=1)
        .unstack(fill_value=0)
    )

    # Previous day totals
    prev_incoming = (
        previous.groupby([type_col, cat_col, "AccountNorm"])["Sum.Incoming"]  # type: ignore
        .sum(min_count=1)
        .unstack(fill_value=0)
    )
    prev_outgoing = (
        previous.groupby([type_col, cat_col, "AccountNorm"])["Sum.Outgoing"]  # type: ignore
        .sum(min_count=1)
        .unstack(fill_value=0)
    )

    # Calculate daily movements (current - previous)
    daily_incoming = curr_incoming.sub(prev_incoming, fill_value=0)
    daily_outgoing = curr_outgoing.sub(prev_outgoing, fill_value=0)

    return daily_incoming, daily_outgoing


def build_cashflow_tables_for_day(raw_json_day: Dict[str, Any], raw_json_prev: Dict[str, Any]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    day = pd.json_normalize(raw_json_day.get("data", []))
    prev = pd.json_normalize(raw_json_prev.get("data", []))

    day = _normalize_accounts(day)
    prev = _normalize_accounts(prev)

    for df in (day, prev):
        for col in [
            "Sum.Incoming",
            "Sum.Outgoing",
            "FinalBalance.Money",
            "StartBalance.Money",
            "CashFlowCategory.HierarchyLevel1",
            "CashFlowCategory.HierarchyLevel2",
            "CashFlowCategory.HierarchyLevel3",
            "CashFlowCategory.Type",
        ]:
            if col not in df.columns:
                df[col] = 0

    cat_col = "CashFlowCategory.HierarchyLevel1"
    type_col = "CashFlowCategory.Type"

    # Start balance at beginning of day: previous day's final balances (no category)
    prev_balances = prev[prev[cat_col].isna()]
    start_bal = (
        prev_balances.groupby("AccountNorm")["FinalBalance.Money"].sum().reindex([RU_ACCOUNT_MAIN, RU_ACCOUNT_TRADES]).fillna(0)
    )

    # Day flows (include all types)
    flows_mask_day = (day[cat_col].notna() if day[cat_col].notna().any() else day[cat_col] != None)
    flows_mask_prev = (prev[cat_col].notna() if prev[cat_col].notna().any() else prev[cat_col] != None)
    flows_day = day[flows_mask_day]
    flows_prev = prev[flows_mask_prev]

    # Calculate daily movements
    incoming, outgoing = calculate_daily_movement(flows_day, flows_prev, "AccountNorm", type_col, cat_col)

    # Map categories to Russian labels and aggregate duplicates
    if not incoming.empty:
        incoming.index = incoming.index.map(lambda x: (x[0], CATEGORY_RU_MAP.get(x[1], str(x[1]))))
        incoming = incoming.groupby(level=[0, 1]).sum()
    if not outgoing.empty:
        outgoing.index = outgoing.index.map(lambda x: (x[0], CATEGORY_RU_MAP.get(x[1], str(x[1]))))
        outgoing = outgoing.groupby(level=[0, 1]).sum()

    # Ensure columns exist
    for frame in (incoming, outgoing):
        for col in [RU_ACCOUNT_MAIN, RU_ACCOUNT_TRADES]:
            if col not in frame.columns:
                frame[col] = 0

    # Calculate end balance using daily movements
    net_by_acc = pd.Series(0, index=[RU_ACCOUNT_MAIN, RU_ACCOUNT_TRADES])
    if not incoming.empty and not outgoing.empty:
        for acc in [RU_ACCOUNT_MAIN, RU_ACCOUNT_TRADES]:
            total_in = incoming.sum()[acc] if acc in incoming.columns else 0
            total_out = outgoing.sum()[acc] if acc in outgoing.columns else 0
            net_by_acc[acc] = total_in - total_out

    end_bal = start_bal + net_by_acc

    # Build result DataFrame
    rows = []
    # Начальные остатки
    rows.append({
        "Тип статьи": "Остаток на начало",
        "Статья ДДС 1-го уровня": "",
        "Статья ДДС 2-го уровня": "",
        "Статья ДДС 3-го уровня": "",
        RU_ACCOUNT_MAIN: float(start_bal.get(RU_ACCOUNT_MAIN, 0)),
        RU_ACCOUNT_TRADES: float(start_bal.get(RU_ACCOUNT_TRADES, 0)),
    })

    # Движения по типам и категориям
    for flow_type in ["OPERATIONAL", "FINANCE"]:
        type_label = "Операционная деятельность" if flow_type == "OPERATIONAL" else "Финансовая деятельность"
        
        # Фильтруем категории для текущего типа
        type_incoming = incoming[incoming.index.get_level_values(0) == flow_type] if not incoming.empty else pd.DataFrame()
        type_outgoing = outgoing[outgoing.index.get_level_values(0) == flow_type] if not outgoing.empty else pd.DataFrame()
        
        if not type_incoming.empty or not type_outgoing.empty:
            # Заголовок раздела
            rows.append({
                "Тип статьи": type_label,
                "Статья ДДС 1-го уровня": "",
                "Статья ДДС 2-го уровня": "",
                "Статья ДДС 3-го уровня": "",
                RU_ACCOUNT_MAIN: 0,
                RU_ACCOUNT_TRADES: 0,
            })

            # Категории текущего типа (корректное объединение множеств)
            cats_in = list(type_incoming.index.get_level_values(1)) if not type_incoming.empty else []
            cats_out = list(type_outgoing.index.get_level_values(1)) if not type_outgoing.empty else []
            categories = sorted(set(cats_in) | set(cats_out))

            for cat in categories:
                # Приход
                if not type_incoming.empty and cat in type_incoming.index.get_level_values(1):
                    cat_incoming = type_incoming.xs((flow_type, cat), level=[0, 1])
                    rows.append({
                        "Тип статьи": type_label,
                        "Статья ДДС 1-го уровня": cat,
                        "Статья ДДС 2-го уровня": "",
                        "Статья ДДС 3-го уровня": "",
                        RU_ACCOUNT_MAIN: float(cat_incoming.get(RU_ACCOUNT_MAIN, 0)),
                        RU_ACCOUNT_TRADES: float(cat_incoming.get(RU_ACCOUNT_TRADES, 0)),
                    })

                # Расход (отрицательные значения)
                if not type_outgoing.empty and cat in type_outgoing.index.get_level_values(1):
                    cat_outgoing = type_outgoing.xs((flow_type, cat), level=[0, 1])
                    rows.append({
                        "Тип статьи": type_label,
                        "Статья ДДС 1-го уровня": cat,
                        "Статья ДДС 2-го уровня": "",
                        "Статья ДДС 3-го уровня": "",
                        RU_ACCOUNT_MAIN: -float(cat_outgoing.get(RU_ACCOUNT_MAIN, 0)),
                        RU_ACCOUNT_TRADES: -float(cat_outgoing.get(RU_ACCOUNT_TRADES, 0)),
                    })

    # Конечные остатки
    rows.append({
        "Тип статьи": "Остаток на конец",
        "Статья ДДС 1-го уровня": "",
        "Статья ДДС 2-го уровня": "",
        "Статья ДДС 3-го уровня": "",
        RU_ACCOUNT_MAIN: float(end_bal.get(RU_ACCOUNT_MAIN, 0)),
        RU_ACCOUNT_TRADES: float(end_bal.get(RU_ACCOUNT_TRADES, 0)),
    })

    # Создаем DataFrame с нужными колонками
    result = pd.DataFrame(rows, columns=[
        "Тип статьи",
        "Статья ДДС 1-го уровня",
        "Статья ДДС 2-го уровня",
        "Статья ДДС 3-го уровня",
        RU_ACCOUNT_MAIN,
        RU_ACCOUNT_TRADES,
    ])

    # Add total column
    result["Итого"] = result[[RU_ACCOUNT_MAIN, RU_ACCOUNT_TRADES]].sum(axis=1)

    return result, day


def build_cashflow_detailed_table_for_day(raw_json_day: Dict[str, Any], raw_json_prev: Dict[str, Any]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    day = pd.json_normalize(raw_json_day.get("data", []))
    prev = pd.json_normalize(raw_json_prev.get("data", []))

    day = _normalize_accounts(day)
    prev = _normalize_accounts(prev)

    for df in (day, prev):
        for col in [
            "Sum.Incoming",
            "Sum.Outgoing",
            "FinalBalance.Money",
            "StartBalance.Money",
            "CashFlowCategory.HierarchyLevel1",
            "CashFlowCategory.HierarchyLevel2",
            "CashFlowCategory.HierarchyLevel3",
            "CashFlowCategory.Type",
        ]:
            if col not in df.columns:
                df[col] = 0

    cat_col = "CashFlowCategory.HierarchyLevel1"
    type_col = "CashFlowCategory.Type"

    prev_balances = prev[prev[cat_col].isna()] if prev[cat_col].isna().any() else prev[prev[cat_col] == None]
    start_bal = (
        prev_balances.groupby("AccountNorm")["FinalBalance.Money"].sum().reindex([RU_ACCOUNT_MAIN, RU_ACCOUNT_TRADES]).fillna(0)
    )

    flows_mask = day[cat_col].notna()
    # В детальном отчёте за день учитываем все типы (OPERATIONAL и FINANCE), чтобы не терялась категория "Займ"
    flows_day = day[flows_mask]

    incoming = (
        flows_day.groupby([type_col, cat_col, "AccountNorm"])  # type: ignore
        ["Sum.Incoming"].sum(min_count=1)
        .unstack(fill_value=0)
    )
    outgoing = (
        flows_day.groupby([type_col, cat_col, "AccountNorm"])  # type: ignore
        ["Sum.Outgoing"].sum(min_count=1)
        .unstack(fill_value=0)
    )

    if not incoming.empty:
        incoming.index = incoming.index.map(lambda x: (x[0], CATEGORY_RU_MAP.get(x[1], str(x[1]))))
        incoming = incoming.groupby(level=[0, 1]).sum()
    if not outgoing.empty:
        outgoing.index = outgoing.index.map(lambda x: (x[0], CATEGORY_RU_MAP.get(x[1], str(x[1]))))
        outgoing = outgoing.groupby(level=[0, 1]).sum()

    for frame in (incoming, outgoing):
        for col in [RU_ACCOUNT_MAIN, RU_ACCOUNT_TRADES]:
            if col not in frame.columns:
                frame[col] = 0

    def _get_vals(frame: pd.DataFrame, cat: str) -> Tuple[float, float]:
        if frame.empty or cat not in frame.index:
            return (0.0, 0.0)
        row = frame.loc[cat]
        return float(row.get(RU_ACCOUNT_MAIN, 0) or 0), float(row.get(RU_ACCOUNT_TRADES, 0) or 0)

    rows = []

    def add_row(type_label: str, l1: str = "", l2: str = "", l3: str = "", inout: str = "", main_val: float = 0.0, trade_val: float = 0.0):
        total = (main_val or 0) + (trade_val or 0)
        rows.append({
            "Тип статьи": type_label,
            "Статья ДДС 1-го уровня": l1,
            "Статья ДДС 2-ого уровня": l2,
            "Статья ДДС 3-ого уровня": l3,
            "Приход/Расход": inout,
            RU_ACCOUNT_MAIN: main_val,
            RU_ACCOUNT_TRADES: trade_val,
            "Итого": total,
        })

    # Start of day
    add_row("Остаток на начало", main_val=float(start_bal.get(RU_ACCOUNT_MAIN, 0)), trade_val=float(start_bal.get(RU_ACCOUNT_TRADES, 0)))

    # Движения по типам и категориям за день: включаем обе категории деятельности
    for flow_type in ["OPERATIONAL", "FINANCE"]:
        type_label = "Операционная деятельность" if flow_type == "OPERATIONAL" else "Финансовая деятельность"

        type_incoming = incoming[incoming.index.get_level_values(0) == flow_type] if not incoming.empty else pd.DataFrame()
        type_outgoing = outgoing[outgoing.index.get_level_values(0) == flow_type] if not outgoing.empty else pd.DataFrame()

        if not type_incoming.empty or not type_outgoing.empty:
            # Заголовок/раздел
            add_row(type_label)

            # Категории текущего типа
            cats_in = list(type_incoming.index.get_level_values(1)) if not type_incoming.empty else []
            cats_out = list(type_outgoing.index.get_level_values(1)) if not type_outgoing.empty else []
            categories = sorted(set(cats_in) | set(cats_out))

            for cat in categories:
                # Приход
                if not type_incoming.empty and cat in type_incoming.index.get_level_values(1):
                    cat_incoming = type_incoming.xs((flow_type, cat), level=[0, 1])
                    add_row(type_label, cat, inout="Приход",
                            main_val=float(cat_incoming.get(RU_ACCOUNT_MAIN, 0)),
                            trade_val=float(cat_incoming.get(RU_ACCOUNT_TRADES, 0)))

                # Расход (отображаем отрицательным)
                if not type_outgoing.empty and cat in type_outgoing.index.get_level_values(1):
                    cat_outgoing = type_outgoing.xs((flow_type, cat), level=[0, 1])
                    add_row(type_label, cat, inout="Расход",
                            main_val=-float(cat_outgoing.get(RU_ACCOUNT_MAIN, 0)),
                            trade_val=-float(cat_outgoing.get(RU_ACCOUNT_TRADES, 0)))

    net_by_acc = (
        flows_day.groupby("AccountNorm").apply(lambda x: (x["Sum.Incoming"].fillna(0) - x["Sum.Outgoing"].fillna(0)).sum())
        .reindex([RU_ACCOUNT_MAIN, RU_ACCOUNT_TRADES])
        .fillna(0)
    )
    end_bal = start_bal + net_by_acc
    add_row("Остаток на конец", main_val=float(end_bal.get(RU_ACCOUNT_MAIN, 0)), trade_val=float(end_bal.get(RU_ACCOUNT_TRADES, 0)))

    detailed = pd.DataFrame(rows, columns=[
        "Тип статьи",
        "Статья ДДС 1-го уровня",
        "Статья ДДС 2-ого уровня",
        "Статья ДДС 3-ого уровня",
        "Приход/Расход",
        RU_ACCOUNT_MAIN,
        RU_ACCOUNT_TRADES,
        "Итого",
    ])

    return detailed, day


def export_to_excel(result: pd.DataFrame, path: str = "cashflow_pivot.xlsx") -> str:
    try:
        result.to_excel(path, sheet_name="Отчет о движении денежных средс", index=False)
        return path
    except PermissionError:
        # Файл может быть открыт в Excel — сохраняем под новым именем с меткой времени
        from datetime import datetime
        root, ext = os.path.splitext(path)
        ext = ext or ".xlsx"
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        alt_path = f"{root}_{ts}{ext}"
        result.to_excel(alt_path, sheet_name="Отчет о движении денежных средс", index=False)
        return alt_path


def dataframe_to_text_table(df: pd.DataFrame, max_rows: int = 30) -> str:
    display_df = df.copy()
    if len(display_df) > max_rows:
        display_df = display_df.head(max_rows)
    def _fmt(v):
        if isinstance(v, (int, float)):
            # Русский формат: пробел как разделитель тысяч и запятая как десятичный
            return f"{v:,.2f}".replace(",", " ").replace(".", ",")
        return str(v)
    formatted = display_df.map(_fmt)
    return formatted.to_string()


# --- Excel mapping: JSON_previous + JSON_current -> Excel cashflow layout ---
def build_excel_cashflow_table(json_prev: Dict[str, Any], json_curr: Dict[str, Any]) -> pd.DataFrame:
    """
    Create Excel-ready table with columns A-R based on mapping rules:
    A: Тип статьи ДДС (mapped from CashFlowCategory.Type)
    B-D: Hierarchy levels
    E-I: Торговые кассы (Start E, Incoming G, Outgoing H, Final I)
    J-N: Главная касса (Start J, Incoming L, Outgoing M, Final N)
    O-R: Итого (Start O, Incoming P, Outgoing Q, Final R)

    Special rows: "Операционная деятельность всего", "Финансовая деятельность всего", "Итого".
    """
    prev = pd.json_normalize(json_prev.get("data", []))
    curr = pd.json_normalize(json_curr.get("data", []))

    prev = _normalize_accounts(prev)
    curr = _normalize_accounts(curr)
    # Ensure columns exist and coerce numerics
    numeric_cols = [
        "Sum.Incoming",
        "Sum.Outgoing",
        "FinalBalance.Money",
        "StartBalance.Money",
    ]
    meta_cols = [
        "CashFlowCategory.HierarchyLevel1",
        "CashFlowCategory.HierarchyLevel2",
        "CashFlowCategory.HierarchyLevel3",
        "CashFlowCategory.Type",
    ]
    for df in (prev, curr):
        for col in numeric_cols + meta_cols:
            if col not in df.columns:
                df[col] = 0.0 if col in numeric_cols else None
        for col in numeric_cols:
            try:
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)
            except Exception:
                # Fallback to zeros if conversion fails
                df[col] = 0.0

    # Diagnostics: total sums to verify non-zero data
    try:
        prev_tot = {c: float(prev[c].sum()) for c in numeric_cols if c in prev.columns}
        curr_tot = {c: float(curr[c].sum()) for c in numeric_cols if c in curr.columns}
        logging.info("[cashflow] prev totals: %s", prev_tot)
        logging.info("[cashflow] curr totals: %s", curr_tot)
    except Exception:
        pass

    # Helper mapping for type
    def type_ru(t: Any) -> str:
        if t == "OPERATIONAL":
            return "Операционная деятельность"
        if t == "FINANCE":
            return "Финансовая деятельность"
        return ""

    # Build key by Account + L1
    def make_key(row: pd.Series) -> Tuple[str, Any]:
        return (str(row.get("AccountNorm", "")), row.get("CashFlowCategory.HierarchyLevel1"))

    # Start balances from previous day: rows where L1 is null -> account totals
    prev_balances = prev[prev["CashFlowCategory.HierarchyLevel1"].isna()] if len(prev) else prev
    start_by_acc = prev_balances.groupby("AccountNorm")["FinalBalance.Money"].sum() if not prev_balances.empty else pd.Series(dtype=float)

    # Current day movements and final balances
    curr_flows = curr[curr["CashFlowCategory.HierarchyLevel1"].notna()] if len(curr) else curr
    # Current day end balances (rows where L1 is null)
    curr_balances = curr[curr["CashFlowCategory.HierarchyLevel1"].isna()] if len(curr) else curr
    end_by_acc = curr_balances.groupby("AccountNorm")["FinalBalance.Money"].sum() if not curr_balances.empty else start_by_acc

    # Aggregate incoming/outgoing per account + L1 + type
    agg_cols = [
        "CashFlowCategory.Type",
        "CashFlowCategory.HierarchyLevel1",
        "CashFlowCategory.HierarchyLevel2",
        "CashFlowCategory.HierarchyLevel3",
        "AccountNorm",
    ]
    if not curr_flows.empty:
        flows_agg = (
            curr_flows.groupby(agg_cols)[["Sum.Incoming", "Sum.Outgoing", "FinalBalance.Money"]]
            .sum()
            .reset_index()
        )
    else:
        flows_agg = pd.DataFrame(columns=agg_cols + ["Sum.Incoming", "Sum.Outgoing", "FinalBalance.Money"])  # empty

    # Collect all unique keys from current flows
    keys = []
    for _, r in flows_agg.iterrows():
        keys.append(make_key(r))
    # Ensure at least totals rows exist
    keys = sorted(set(keys))

    rows = []

    def add_row(type_label: str, l1: Any, l2: Any, l3: Any,
                start_trade: float, in_trade: float, out_trade: float, fin_trade: float,
                start_main: float, in_main: float, out_main: float, fin_main: float):
        # Totals O-R
        start_total = (start_trade or 0) + (start_main or 0)
        in_total = (in_trade or 0) + (in_main or 0)
        out_total = (out_trade or 0) + (out_main or 0)
        fin_total = (fin_trade or 0) + (fin_main or 0)
        rows.append({
            "Тип статьи ДДС": type_label,
            "Статья ДДС 1-го уровня": CATEGORY_RU_MAP.get(l1, l1) if l1 is not None else "",
            "Статья ДДС 2-го уровня": l2 or "",
            "Статья ДДС 3-го уровня": l3 or "",
            # E-I: Главная касса (keep F blank)
            "E": float(start_main or 0),
            "F": "",
            "G": float(in_main or 0),
            "H": float(out_main or 0),
            "I": float(fin_main or 0),
            # J-N: Торговые кассы (keep K blank)
            "J": float(start_trade or 0),
            "K": "",
            "L": float(in_trade or 0),
            "M": float(out_trade or 0),
            "N": float(fin_trade or 0),
            # O-R: Итого
            "O": float(start_total),
            "P": float(in_total),
            "Q": float(out_total),
            "R": float(fin_total),
        })

    # Row 9: Общие остатки по кассам
    # E9/J9: FinalBalance (prev, L1 is null) for Main/Trades; I9/N9: FinalBalance (curr, L1 is null)
    try:
        start_trade_val = float(start_by_acc.get(RU_ACCOUNT_TRADES, 0) or 0)
    except Exception:
        start_trade_val = 0.0
    try:
        start_main_val = float(start_by_acc.get(RU_ACCOUNT_MAIN, 0) or 0)
    except Exception:
        start_main_val = 0.0
    try:
        end_trade_val = float(end_by_acc.get(RU_ACCOUNT_TRADES, 0) or 0)
    except Exception:
        end_trade_val = 0.0
    try:
        end_main_val = float(end_by_acc.get(RU_ACCOUNT_MAIN, 0) or 0)
    except Exception:
        end_main_val = 0.0

    add_row(
        "Общие остатки по кассам", None, None, None,
        start_trade_val, 0.0, 0.0, end_trade_val,
        start_main_val, 0.0, 0.0, end_main_val,
    )

    # --- Explicit rows mapping per specification ---
    # Helper: build synonyms for a Russian category label
    def l1_synonyms(ru_label: str) -> set:
        syn = {ru_label}
        try:
            syn |= {k for k, v in CATEGORY_RU_MAP.items() if v == ru_label}
        except Exception:
            pass
        return syn

    # Prepare per-account, per-category aggregations from prev/curr
    prev_cat = prev[prev["CashFlowCategory.HierarchyLevel1"].notna()]
    curr_cat = curr[curr["CashFlowCategory.HierarchyLevel1"].notna()]
    prev_fb = prev_cat.groupby(["AccountNorm", "CashFlowCategory.HierarchyLevel1"])  # type: ignore
    prev_fb = prev_fb["FinalBalance.Money"].sum().to_dict() if not prev_cat.empty else {}
    curr_fb = curr_cat.groupby(["AccountNorm", "CashFlowCategory.HierarchyLevel1"])  # type: ignore
    curr_fb = curr_fb["FinalBalance.Money"].sum().to_dict() if not curr_cat.empty else {}
    curr_in = curr_cat.groupby(["AccountNorm", "CashFlowCategory.HierarchyLevel1"])  # type: ignore
    curr_in = curr_in["Sum.Incoming"].sum().to_dict() if not curr_cat.empty else {}
    curr_out = curr_cat.groupby(["AccountNorm", "CashFlowCategory.HierarchyLevel1"])  # type: ignore
    curr_out = curr_out["Sum.Outgoing"].sum().to_dict() if not curr_cat.empty else {}

    def get_val(dct: Dict[Tuple[str, Any], float], acc: str, ru_label: str) -> float:
        total = 0.0
        for name in l1_synonyms(ru_label):
            total += float(dct.get((acc, name), 0.0) or 0.0)
        return total

    # Row 10: Операционная деятельность / Внутреннее перемещение
    l1 = "Внутреннее перемещение"
    add_row(
        "Операционная деятельность", l1, None, None,
        # trades
        get_val(prev_fb, RU_ACCOUNT_TRADES, l1), 0.0, get_val(curr_out, RU_ACCOUNT_TRADES, l1), get_val(curr_fb, RU_ACCOUNT_TRADES, l1),
        # main
        get_val(prev_fb, RU_ACCOUNT_MAIN, l1), get_val(curr_in, RU_ACCOUNT_MAIN, l1), 0.0, get_val(curr_fb, RU_ACCOUNT_MAIN, l1),
    )

    # Row 11: Выручка (operational, trades only)
    l1 = "Выручка"
    add_row(
        "", l1, None, None,
        # trades
        get_val(prev_fb, RU_ACCOUNT_TRADES, l1), get_val(curr_in, RU_ACCOUNT_TRADES, l1), 0.0, get_val(curr_fb, RU_ACCOUNT_TRADES, l1),
        # main (zeros)
        0.0, 0.0, 0.0, 0.0,
    )

    # Row 12: Оплата накладных (both accounts, incoming/outgoing zero)
    l1 = "Оплата накладных"
    add_row(
        "", l1, None, None,
        # trades
        get_val(prev_fb, RU_ACCOUNT_TRADES, l1), 0.0, 0.0, get_val(curr_fb, RU_ACCOUNT_TRADES, l1),
        # main
        get_val(prev_fb, RU_ACCOUNT_MAIN, l1), 0.0, 0.0, get_val(curr_fb, RU_ACCOUNT_MAIN, l1),
    )

    # Row 13: Оплата труда (main only)
    l1 = "Оплата труда"
    add_row(
        "", l1, None, None,
        # trades (zeros)
        0.0, 0.0, 0.0, 0.0,
        # main
        get_val(prev_fb, RU_ACCOUNT_MAIN, l1), 0.0, 0.0, get_val(curr_fb, RU_ACCOUNT_MAIN, l1),
    )

    # Row 14: Подотчет (both accounts; H from main outgoing, M=0)
    l1 = "Подотчет"
    add_row(
        "", l1, None, None,
        # trades
        get_val(prev_fb, RU_ACCOUNT_TRADES, l1), 0.0, 0.0, get_val(curr_fb, RU_ACCOUNT_TRADES, l1),
        # main
        get_val(prev_fb, RU_ACCOUNT_MAIN, l1), 0.0, get_val(curr_out, RU_ACCOUNT_MAIN, l1), get_val(curr_fb, RU_ACCOUNT_MAIN, l1),
    )

    # Row 15: Предоплата (trades only)
    l1 = "Предоплата"
    add_row(
        "", l1, None, None,
        # trades
        get_val(prev_fb, RU_ACCOUNT_TRADES, l1), 0.0, 0.0, get_val(curr_fb, RU_ACCOUNT_TRADES, l1),
        # main (zeros)
        0.0, 0.0, 0.0, 0.0,
    )

    # Row 16: Операционная деятельность всего (sum rows 10-15)
    op_rows = rows[-6:]
    def sum_list(col: str) -> float:
        try:
            return float(pd.DataFrame(op_rows)[col].sum())
        except Exception:
            return 0.0
    add_row(
        "Операционная деятельность всего", None, None, None,
        # trades sums
        sum_list("J"), sum_list("L"), sum_list("M"), sum_list("N"),
        # main sums
        sum_list("E"), sum_list("G"), sum_list("H"), sum_list("I"),
    )

    # Row 17: Финансовая деятельность / Займ (main only)
    l1 = "Займ"
    add_row(
        "Финансовая деятельность", l1, None, None,
        # trades (zeros)
        0.0, 0.0, 0.0, 0.0,
        # main
        get_val(prev_fb, RU_ACCOUNT_MAIN, l1), 0.0, 0.0, get_val(curr_fb, RU_ACCOUNT_MAIN, l1),
    )

    # Row 18: Итого (row9 + row16 + row17)
    # capture rows by index positions based on appended order
    r9 = rows[0]
    r16 = rows[-2]
    r17 = rows[-1]
    add_row(
        "Итого", None, None, None,
        # trades
        float(r9.get("J", 0) or 0) + float(r16.get("J", 0) or 0) + float(r17.get("J", 0) or 0),
        float(r9.get("L", 0) or 0) + float(r16.get("L", 0) or 0) + float(r17.get("L", 0) or 0),
        float(r9.get("M", 0) or 0) + float(r16.get("M", 0) or 0) + float(r17.get("M", 0) or 0),
        float(r9.get("N", 0) or 0) + float(r16.get("N", 0) or 0) + float(r17.get("N", 0) or 0),
        # main
        float(r9.get("E", 0) or 0) + float(r16.get("E", 0) or 0) + float(r17.get("E", 0) or 0),
        float(r9.get("G", 0) or 0) + float(r16.get("G", 0) or 0) + float(r17.get("G", 0) or 0),
        float(r9.get("H", 0) or 0) + float(r16.get("H", 0) or 0) + float(r17.get("H", 0) or 0),
        float(r9.get("I", 0) or 0) + float(r16.get("I", 0) or 0) + float(r17.get("I", 0) or 0),
    )

    # Final DataFrame in display order: A-D then E-R
    table = pd.DataFrame(rows, columns=[
        "Тип статьи ДДС",
        "Статья ДДС 1-го уровня",
        "Статья ДДС 2-го уровня",
        "Статья ДДС 3-го уровня",
        "E", "F", "G", "H", "I",
        "J", "K", "L", "M", "N",
        "O", "P", "Q", "R",
    ])

    return table


def export_excel_cashflow(json_prev: Dict[str, Any], json_curr: Dict[str, Any], date_caption: str, path: str | None = None) -> str:
    """
    Export mapped cashflow to Excel with header (restaurant name + date) and table.
    """
    table = build_excel_cashflow_table(json_prev, json_curr)
    # Format date to dd.mm.yyyy if ISO provided
    try:
        from datetime import datetime
        date_caption_fmt = datetime.fromisoformat(date_caption).strftime("%d.%m.%Y")
    except Exception:
        date_caption_fmt = str(date_caption)

    # If path not provided, default to date-based filename like 09.10.2025.xlsx
    if path is None or not str(path).strip():
        path = f"{date_caption_fmt}.xlsx"

    # Write using openpyxl to build multi-row headers
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet

    wb = Workbook()
    ws: Worksheet = wb.active
    ws.title = "Отчет о движении денежных средс"

    # Fixed headers:
    # B1: Report title
    ws.cell(row=1, column=2, value="Отчет о движении денежных средств1")
    # B2: Restaurant name
    ws.cell(row=2, column=2, value="Название ресторана: Таврика local cafe")
    # B3: Date (from provided caption)
    ws.cell(row=3, column=2, value=f"Дата: {date_caption_fmt}")

    # Row 5: Single header in E5
    ws.cell(row=5, column=5, value="Счет")
    # Row 6: Group headers with merges (E:I, J:N, O:R)
    ws.cell(row=6, column=5, value="Главная касса")
    ws.merge_cells(start_row=6, start_column=5, end_row=6, end_column=9)
    ws.cell(row=6, column=10, value="Торговые кассы")
    ws.merge_cells(start_row=6, start_column=10, end_row=6, end_column=14)
    ws.cell(row=6, column=15, value="Итого")
    ws.merge_cells(start_row=6, start_column=15, end_row=6, end_column=18)

    # Row 8: Column headers
    headers_a_d = [
        "Тип статьи ДДС",
        "Статья ДДС 1-го уровня",
        "Статья ДДС 2-го уровня",
        "Статья ДДС 3-го уровня",
    ]
    for idx, h in enumerate(headers_a_d, start=1):
        ws.cell(row=8, column=idx, value=h)
    # Row 8: Sub-headers for accounts (leave F8 and K8 blank)
    # E-I: Главная касса
    ws.cell(row=8, column=5, value="Начальный денежный остаток, р.")  # E8
    ws.cell(row=8, column=7, value="Сумма прихода, р.")               # G8
    ws.cell(row=8, column=8, value="Сумма расхода, р.")               # H8
    ws.cell(row=8, column=9, value="Конечный денежный остаток, р.")   # I8
    # J-N: Торговые кассы
    ws.cell(row=8, column=10, value="Начальный денежный остаток, р.") # J8
    ws.cell(row=8, column=12, value="Сумма прихода, р.")              # L8
    ws.cell(row=8, column=13, value="Сумма расхода, р.")              # M8
    ws.cell(row=8, column=14, value="Конечный денежный остаток, р.")  # N8
    # O-R: Итого
    ws.cell(row=8, column=15, value="Начальный денежный остаток, р.") # O8
    ws.cell(row=8, column=16, value="Сумма прихода, р.")              # P8
    ws.cell(row=8, column=17, value="Сумма расхода, р.")              # Q8
    ws.cell(row=8, column=18, value="Конечный денежный остаток, р.")  # R8

    # Data rows start at row 9 (after headers)
    start_row = 9
    for r_idx, row in enumerate(table.to_dict(orient="records"), start=start_row):
        t = str(row.get("Тип статьи ДДС", "")).strip()
        # A-D
        ws.cell(row=r_idx, column=1, value=row.get("Тип статьи ДДС", ""))
        ws.cell(row=r_idx, column=2, value=row.get("Статья ДДС 1-го уровня", ""))
        ws.cell(row=r_idx, column=3, value=row.get("Статья ДДС 2-го уровня", ""))
        ws.cell(row=r_idx, column=4, value=row.get("Статья ДДС 3-го уровня", ""))
        # E-R numeric cells (contiguous mapping without separators)
        for col_idx, key in enumerate(["E","F","G","H","I","J","K","L","M","N","O","P","Q","R"], start=5):
            ws.cell(row=r_idx, column=col_idx, value=row.get(key, ""))

    # Save workbook with fallback if the target file is locked/open
    try:
        wb.save(path)
        return path
    except PermissionError:
        from datetime import datetime
        root, ext = os.path.splitext(path)
        ext = ext or ".xlsx"
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        alt_path = f"{root}_{ts}{ext}"
        wb.save(alt_path)
        return alt_path



def build_full_cashflow_tree(table: pd.DataFrame, date_str: str) -> str:
    """
    Преобразует cashflow table в текстовое дерево с ├── / └── и возвращает как строку.
    """
    cash_names = ["Главная касса", "Торговые кассы"]
    cash_columns = {
        "Главная касса": ["E", "G", "H", "I"],
        "Торговые кассы": ["J", "L", "M", "N"],
        "Итого": ["O", "P", "Q", "R"]
    }

    lines = []

    # Рекурсивная функция для обхода dict и генерации текста
    def tree_to_text(node: dict, prefix: str = "") -> list[str]:
        text_lines = []

        if "Итого" in node:
            keys = ["Итого"]
        else:
            keys = list(node.keys())

        for i, key in enumerate(keys):
            is_last = (i == len(keys) - 1)
            pointer = "└──" if is_last else "├──"
            val = node[key]
            if isinstance(val, dict):
                text_lines.append(f"{prefix}{pointer} {key}")
                child_prefix = prefix + ("    " if is_last else "│   ")
                text_lines.extend(tree_to_text(val, child_prefix))
            else:
                text_lines.append(f"{prefix}{pointer} {key}: {val:,.2f}")
        return text_lines

    # Строим дерево для каждой кассы
    def build_tree_for_cash(cash_name: str) -> dict:
        tree = {}
        for _, row in table.iterrows():
            type_label = row.get("Тип статьи ДДС", "").strip()
            l1 = row.get("Статья ДДС 1-го уровня", "").strip()
            l2 = row.get("Статья ДДС 2-го уровня", "").strip()
            l3 = row.get("Статья ДДС 3-го уровня", "").strip()

            node_lvl = tree
            if type_label:
                node_lvl = node_lvl.setdefault(type_label, {})
            if l1:
                node_lvl = node_lvl.setdefault(l1, {})
            if l2:
                node_lvl = node_lvl.setdefault(l2, {})
            if l3:
                node_lvl = node_lvl.setdefault(l3, {})

            for col_name, label in zip(cash_columns[cash_name],
                                       ["Начальный денежный остаток, р.",
                                        "Сумма прихода, р.",
                                        "Сумма расхода, р.",
                                        "Конечный денежный остаток, р."]):
                node_lvl[label] = row.get(col_name, 0.0)
        return tree

    # Собираем строки
    lines.append(f"Дата: {date_str}")
    for cash_name in cash_names:
        lines.append(cash_name)
        cash_tree = build_tree_for_cash(cash_name)
        lines.extend(tree_to_text(cash_tree, prefix="    "))

    return "\n".join(lines)
